#!/usr/bin/env python3
"""leak_csv.py -- shared scaffold for classifying Maltego D4 leaked-record CSV
exports against a case's identity anchors.

Consolidated 2026-07-21 from two independently-written near-duplicates:
  case-001-example/investigation/evidence/scripts/reformat_csv.py
  case-001-example/investigation/evidence/scripts/reformat_csv.py
Both reinvented the same ~250-line scaffold (Section 1/2/3 boundary detection,
csv.DictReader parsing, a colored multi-tab Excel writer, cleaned/filtered-out
CSV writers) with only the identity anchors and tier-classification rules
actually differing per case. This module is that shared scaffold. A case
supplies a small config (see templates/case_classifier_template.py) instead of
rewriting the whole thing.

Those two origin scripts are NOT retroactively migrated to this module --
they were already delivered to clients and stay as-is, unmodified.

What THIS module does NOT do (stays case-specific, in the case's own script):
  - The actual classify_row() identity logic (SSN/DOB/address/email anchors,
    tier names, precedence rules) -- every case's anchors and edge cases differ.
  - Report generation (generate-report.js) -- bespoke narrative per subject.

Maltego export structure (both origin cases confirmed this shape):
  Section 1: main data records (2-row header in some exports: display names
             then technical field names; single-row header in others -- this
             module tries DictReader directly first, and only skips a second
             header row if the parsed fieldnames don't look like real columns)
  Section 2: query-parameter echo entities (Alias, EmailAddress, Home,
             IdentificationNumber, Person, PhoneNumber, Phrase, LinkedIn
             affiliation, etc.) -- the search inputs, not results.
  Section 3: Maltego graph link data (LinkID, SourceEntityID, TargetEntityID, ...)
Only Section 1 is processed. Rows can contain embedded newlines in quoted
fields (e.g. "Other Fields"), so physical line count != logical row count --
csv.DictReader is required, never naive line-splitting.
"""

import csv
import io
import re
import sys
from pathlib import Path

# Default Maltego section-2/3 boundary markers seen across both origin exports.
# A case can pass its own tuple if a new export uses different header text --
# check the raw file for lines starting with "maltego." or an entity-id header
# before assuming these defaults are wrong.
DEFAULT_SECTION_BOUNDARY_MARKERS = (
    "maltego.",
    "entityid,complexquery",
    "linkid,sourceentityid",
    "email address,is partial",
    "email,is_partial",
    "source entity id,target entity id",
)


def norm_digits(raw):
    return re.sub(r"\D", "", raw or "")


def norm_dob(raw):
    """Normalize a birthdate string to YYYY-MM-DD, or None if unparseable.
    Handles YYYY-M-D, YYYY/M/D, M-D-YYYY variants (the formats actually seen
    in both origin exports)."""
    raw = (raw or "").strip()
    if not raw:
        return None
    raw2 = raw.replace("/", "-")
    m = re.match(r"^(\d{4})-(\d{1,2})-(\d{1,2})$", raw2)
    if m:
        y, mo, d = m.groups()
        return f"{y}-{int(mo):02d}-{int(d):02d}"
    m = re.match(r"^(\d{1,2})-(\d{1,2})-(\d{4})$", raw2)
    if m:
        mo, d, y = m.groups()
        return f"{y}-{int(mo):02d}-{int(d):02d}"
    return None


def dob_month_day_matches(raw, target_dob):
    """True if raw's month/day matches target_dob's month/day, regardless of
    year (catches anomalous-year data-entry artifacts, e.g. year 1904 or 2099
    with the right month/day -- seen in the Goodgame origin case)."""
    dob = norm_dob(raw)
    if not dob:
        return False
    return dob[5:] == target_dob[5:]


def addr_hits(addr, confirmed_addr_tokens):
    a = (addr or "").lower()
    return any(t in a for t in confirmed_addr_tokens)


def emails_in_field(raw):
    raw = (raw or "").lower()
    return [e.strip() for e in re.split(r"[\s,;\n]+", raw) if "@" in e]


def detect_section1_end(lines, boundary_markers=DEFAULT_SECTION_BOUNDARY_MARKERS):
    for i, line in enumerate(lines):
        if i < 2:
            continue
        stripped = line.strip().lower()
        if stripped.startswith(boundary_markers):
            cutoff = i
            while cutoff > 0 and lines[cutoff - 1].strip() == "":
                cutoff -= 1
            return cutoff
    return len(lines)


def read_section1(csv_path, boundary_markers=DEFAULT_SECTION_BOUNDARY_MARKERS,
                   entity_id_column_prefix="district4."):
    """Parse Section 1 of a Maltego CSV export. Returns (fieldnames, rows).
    Renames a leading 'district4.<EntityType>' column to 'EntityID' for
    output clarity (seen in both origin exports' first column)."""
    csv_path = Path(csv_path)
    raw = csv_path.read_text(encoding="utf-8", errors="replace")
    lines = raw.splitlines(keepends=True)
    cutoff = detect_section1_end(lines, boundary_markers)
    section1_text = "".join(lines[:cutoff])
    reader = csv.DictReader(io.StringIO(section1_text))
    rows = list(reader)
    fieldnames = reader.fieldnames or []
    if fieldnames and fieldnames[0].startswith(entity_id_column_prefix):
        fieldnames = ["EntityID"] + fieldnames[1:]
    return fieldnames, rows


def build_output_columns(fieldnames, drop_columns=None):
    drop_columns = {c.lower() for c in (drop_columns or set())}
    return [f for f in fieldnames if f and f.lower() not in drop_columns]


def write_csv(path, classified_rows, output_cols, include_tier=True, include_reason=True):
    """classified_rows: list of (row_dict, tier, reason) tuples."""
    extra = []
    if include_tier:
        extra.append("_Tier")
    if include_reason:
        extra.append("_Reason")
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(output_cols + extra)
        for row, tier, reason in classified_rows:
            values = [(row.get(c) or "").strip().replace("\n", " | ") for c in output_cols]
            if include_tier:
                values.append(tier)
            if include_reason:
                values.append(reason)
            writer.writerow(values)


def write_excel(classified_rows, output_path, output_cols, tier_config, text_columns=None):
    """classified_rows: list of (row_dict, tier, reason) tuples.
    tier_config: ordered list of dicts, each {key, sheet_title, fill_color}.
    Only tiers present in tier_config get a sheet (and in that order) --
    a tier not listed (e.g. a case's 'unrelated' tier) is silently excluded
    from the Excel, matching both origin cases' pattern of keeping fully
    unrelated noise out of the client-facing spreadsheet."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("ERROR: openpyxl not installed. Run: pip3 install openpyxl", file=sys.stderr)
        sys.exit(1)

    text_columns = {c.lower() for c in (text_columns or
                    ["Userid", "Ssn", "Creditcard", "Dl License", "Vin", "Facebook Userid"])}

    wb = openpyxl.Workbook()
    HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)

    def write_sheet(ws, rows_with_reason, fill_color):
        fill = PatternFill("solid", fgColor=fill_color)
        cols = output_cols + ["_Reason"]
        for col_idx, col_name in enumerate(cols, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(wrap_text=False)
        for row_idx, (row, tier, reason) in enumerate(rows_with_reason, 2):
            for col_idx, col_name in enumerate(output_cols, 1):
                val = (row.get(col_name) or "").strip().replace("\n", " | ")
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.fill = fill
                cell.alignment = Alignment(wrap_text=False)
                if col_name.lower() in text_columns:
                    cell.number_format = "@"
            cell = ws.cell(row=row_idx, column=len(output_cols) + 1, value=reason)
            cell.fill = fill
        for col_idx, col_name in enumerate(cols, 1):
            max_len = len(col_name)
            for row_data, _, _ in rows_with_reason:
                val = str(row_data.get(col_name) or "")
                max_len = max(max_len, min(len(val), 60))
            ws.column_dimensions[get_column_letter(col_idx)].width = max(10, max_len + 2)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    first = True
    for tier_def in tier_config:
        rows_for_tier = [r for r in classified_rows if r[1] == tier_def["key"]]
        ws = wb.active if first else wb.create_sheet()
        ws.title = tier_def["sheet_title"]
        write_sheet(ws, rows_for_tier, tier_def["fill_color"])
        first = False

    wb.save(output_path)


def read_active_case(workspace):
    pointer = Path(workspace) / ".active-case"
    if not pointer.is_file():
        return ""
    return pointer.read_text().strip()


def resolve_paths(workspace, case, input_override=None, output_dir_override=None):
    workspace = Path(workspace)
    if not case:
        case = read_active_case(workspace)
    if not case:
        print("ERROR: no --case given and no .active-case set", file=sys.stderr)
        sys.exit(1)

    case_dir = workspace / "investigations" / case
    if input_override:
        csv_path = Path(input_override)
    else:
        intake_dir = case_dir / "investigation" / "intake"
        csv_files = list(intake_dir.glob("*.csv"))
        if not csv_files:
            print(f"ERROR: no CSV files found in {intake_dir}", file=sys.stderr)
            sys.exit(1)
        csv_path = csv_files[0]

    output_dir = Path(output_dir_override) if output_dir_override else case_dir / "output"
    output_dir.mkdir(parents=True, exist_ok=True)
    return case, csv_path, output_dir


def run(workspace, case_arg, input_arg, output_dir_arg, classify_fn, tier_config,
        cleaned_tiers, filtered_out_tiers, output_basename,
        drop_columns=None, boundary_markers=DEFAULT_SECTION_BOUNDARY_MARKERS,
        entity_id_column_prefix="district4.", text_columns=None):
    """Orchestrates the full pipeline. A case's script calls this with its own
    classify_fn(row) -> (tier_key, reason) and tier_config (see template)."""
    case, csv_path, output_dir = resolve_paths(workspace, case_arg, input_arg, output_dir_arg)
    print(f"Case:   {case}")
    print(f"Input:  {csv_path}")
    print(f"Output: {output_dir}")

    fieldnames, rows = read_section1(csv_path, boundary_markers, entity_id_column_prefix)
    print(f"Section 1 rows parsed: {len(rows)}")
    output_cols = build_output_columns(fieldnames, drop_columns)

    classified = [(row, *classify_fn(row)) for row in rows]

    counts = {}
    for _, tier, _ in classified:
        counts[tier] = counts.get(tier, 0) + 1
    print("Classification counts:", counts)

    cleaned_rows = [r for r in classified if r[1] in cleaned_tiers]
    filtered_out_rows = [r for r in classified if r[1] in filtered_out_tiers]

    write_csv(output_dir / f"{output_basename}_cleaned.csv", cleaned_rows, output_cols)
    write_csv(output_dir / f"{output_basename}_filtered_out.csv", filtered_out_rows, output_cols)
    print(f"Wrote {len(cleaned_rows)} rows -> {output_basename}_cleaned.csv")
    print(f"Wrote {len(filtered_out_rows)} rows -> {output_basename}_filtered_out.csv")

    excel_tier_keys = {t["key"] for t in tier_config}
    excel_rows = [r for r in classified if r[1] in excel_tier_keys]
    xlsx_path = output_dir / f"{output_basename}_Cleaned_Records.xlsx"
    write_excel(excel_rows, xlsx_path, output_cols, tier_config, text_columns)
    print(f"Wrote {xlsx_path.name}")

    return classified
